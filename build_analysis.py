import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── Load and prep ─────────────────────────────────────────────────────────────
df = pd.read_excel("CU Spend Data_v1-2.xlsx")
pkg = df[df['AIC Category 2'] == 'PACKAGING'].copy()

classifications = {
    125: 'STRETCH WRAP', 209: 'END BAGS', 383: 'PACKAGING LABELS',
    541: 'STRAP TAPE',   691: 'PACKAGING LABELS', 744: 'STRETCH WRAP',
    753: 'STRAP TAPE',   801: 'LAYER PADS',        962: 'STRETCH WRAP',
}
for idx, cat in classifications.items():
    pkg.at[idx, 'AIC Category 3'] = cat

pkg = pkg.reset_index(drop=True)
total_spend = pkg['AIC Spend'].sum()

# ── Style helpers ─────────────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FONT = Font(bold=True, size=11)
CURRENCY_FMT = '"$"#,##0.00'
PERCENT_FMT  = '0.00%'
CENTER       = Alignment(horizontal='center')

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

def freeze_and_fit(ws):
    ws.freeze_panes = "A2"
    autofit(ws)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Raw Data (Cleaned)
# ════════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Raw Data (Cleaned)"

headers = list(pkg.columns)
ws1.append(headers)
style_header_row(ws1, 1, len(headers))

spend_col = headers.index('AIC Spend') + 1

for r_idx, row in pkg.iterrows():
    ws1.append(list(row))
    ws1.cell(row=r_idx + 2, column=spend_col).number_format = CURRENCY_FMT

freeze_and_fit(ws1)

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Spend Summary
# ════════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Spend Summary")

ws2.append(["PACKAGING SPEND SUMMARY"])
ws2['A1'].font = Font(bold=True, size=14)
ws2.append([])
ws2.append(["Total Category Spend", total_spend])
ws2['A3'].font = SUBHEAD_FONT
ws2['B3'].number_format = CURRENCY_FMT
ws2['B3'].font = Font(bold=True)
ws2.append([])

ws2.append(["Sub-Category (Category 3)", "Spend ($)", "% of Total"])
style_header_row(ws2, 5, 3)

cat3_summary = pkg.groupby('AIC Category 3')['AIC Spend'].sum().sort_values(ascending=False).reset_index()
for _, row in cat3_summary.iterrows():
    pct = row['AIC Spend'] / total_spend
    ws2.append([row['AIC Category 3'], row['AIC Spend'], pct])
    ws2.cell(row=ws2.max_row, column=2).number_format = CURRENCY_FMT
    ws2.cell(row=ws2.max_row, column=3).number_format = PERCENT_FMT

ws2.append([])

bu_row = ws2.max_row + 1
ws2.append(["Business Unit / Region", "Spend ($)", "% of Total"])
style_header_row(ws2, bu_row, 3)

bu_summary = pkg.groupby('Business Unit')['AIC Spend'].sum().sort_values(ascending=False).reset_index()
for _, row in bu_summary.iterrows():
    pct = row['AIC Spend'] / total_spend
    ws2.append([row['Business Unit'], row['AIC Spend'], pct])
    ws2.cell(row=ws2.max_row, column=2).number_format = CURRENCY_FMT
    ws2.cell(row=ws2.max_row, column=3).number_format = PERCENT_FMT

freeze_and_fit(ws2)

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Supplier Analysis
# ════════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Supplier Analysis")

supplier_summary = pkg.groupby('AIC Supplier Name')['AIC Spend'].sum().sort_values(ascending=False).reset_index()
n_suppliers = len(supplier_summary)

ws3.append(["SUPPLIER ANALYSIS"])
ws3['A1'].font = Font(bold=True, size=14)
ws3.append([])
ws3.append(["Total Unique Suppliers", n_suppliers])
ws3['A3'].font = SUBHEAD_FONT
ws3.append([])

ws3.append(["Top 5 Suppliers by Spend"])
ws3.cell(row=ws3.max_row, column=1).font = SUBHEAD_FONT
ws3.append(["Supplier", "Spend ($)", "% of Total"])
style_header_row(ws3, ws3.max_row, 3)

for _, row in supplier_summary.head(5).iterrows():
    pct = row['AIC Spend'] / total_spend
    ws3.append([row['AIC Supplier Name'], row['AIC Spend'], pct])
    ws3.cell(row=ws3.max_row, column=2).number_format = CURRENCY_FMT
    ws3.cell(row=ws3.max_row, column=3).number_format = PERCENT_FMT

ws3.append([])

ws3.append(["Pareto Table — Cumulative Supplier Spend"])
ws3.cell(row=ws3.max_row, column=1).font = SUBHEAD_FONT
ws3.append(["Rank", "Supplier", "Spend ($)", "% of Total", "Cumulative %"])
style_header_row(ws3, ws3.max_row, 5)

cumulative = 0
for rank, (_, row) in enumerate(supplier_summary.iterrows(), 1):
    pct = row['AIC Spend'] / total_spend
    cumulative += pct
    ws3.append([rank, row['AIC Supplier Name'], row['AIC Spend'], pct, cumulative])
    r = ws3.max_row
    ws3.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws3.cell(row=r, column=4).number_format = PERCENT_FMT
    ws3.cell(row=r, column=5).number_format = PERCENT_FMT

freeze_and_fit(ws3)

# ════════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Pareto Chart
# ════════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Pareto Chart")

ws4.append(["Supplier", "Spend ($)", "Cumulative %"])
style_header_row(ws4, 1, 3)

cumulative = 0
for _, row in supplier_summary.iterrows():
    pct = row['AIC Spend'] / total_spend
    cumulative += pct
    ws4.append([row['AIC Supplier Name'], row['AIC Spend'], cumulative])
    r = ws4.max_row
    ws4.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=3).number_format = PERCENT_FMT

n_rows = len(supplier_summary)

# Bar chart
bar = BarChart()
bar.type = "col"
bar.title = "Supplier Spend Pareto"
bar.y_axis.title = "Spend ($)"
bar.x_axis.title = "Supplier"
bar.style = 10
bar.width = 28
bar.height = 16

data_ref = Reference(ws4, min_col=2, min_row=1, max_row=n_rows + 1)
cats_ref  = Reference(ws4, min_col=1, min_row=2, max_row=n_rows + 1)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.series[0].graphicalProperties.solidFill = "1F4E79"

# Line chart (cumulative %)
line = LineChart()
line.y_axis.axId = 200
line.y_axis.title = "Cumulative %"
line.y_axis.numFmt = "0%"
line.y_axis.crosses = "max"

cum_ref = Reference(ws4, min_col=3, min_row=1, max_row=n_rows + 1)
line.add_data(cum_ref, titles_from_data=True)
line.series[0].graphicalProperties.line.solidFill = "FF0000"
line.series[0].graphicalProperties.line.width = 20000

bar += line
ws4.add_chart(bar, "E2")

freeze_and_fit(ws4)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save("Spend_Analysis_Output.xlsx")
print("Spend_Analysis_Output.xlsx saved successfully!")
print("\nSheets created:")
for ws in wb.worksheets:
    print(f"  - {ws.title}")

print(f"\nKey stats:")
print(f"  Total PACKAGING spend: ${total_spend:,.2f}")
print(f"  Unique suppliers: {n_suppliers}")
print(f"  Sub-categories: {len(cat3_summary)}")
